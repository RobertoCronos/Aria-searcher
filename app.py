import streamlit as st
import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
import json, re, io
from datetime import datetime

# ── Page config ──────────────────────────────────────────────────────
st.set_page_config(
    page_title="ARIA Membresías",
    page_icon="🏆",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── CSS ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Mono:wght@400;500&family=Sora:wght@300;400;500;600&display=swap');
html, body, [class*="css"] { font-family: 'Sora', sans-serif; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 2rem; padding-bottom: 3rem; max-width: 860px; }

.aria-header {
    background: linear-gradient(160deg, #E8C800 0%, #B88A00 35%, #F5F0D0 70%, #FAFAF5 100%);
    border-radius: 16px; padding: 36px 36px 32px; text-align: center;
    margin-bottom: 24px; border: none;
    position: relative; overflow: hidden;
}
.aria-eyebrow {
    display: inline-flex; align-items: center; gap: 6px;
    background: rgba(255,255,255,0.25); border: 1px solid rgba(255,255,255,0.45);
    border-radius: 100px; padding: 5px 16px;
    font-size: 10px; letter-spacing: 2px; color: rgba(50,30,0,0.8);
    text-transform: uppercase; font-weight: 600; margin-bottom: 16px;
}
.aria-title { font-family: 'DM Serif Display', serif; font-size: 2.4rem; color: #1A1200; margin: 0; }
.aria-title em { font-style: italic; color: #7A4F00; }
.aria-sub { font-size: 0.82rem; color: rgba(50,30,0,0.58); margin-top: 6px; }

.section-label {
    font-size: 0.68rem; letter-spacing: 2.5px; text-transform: uppercase;
    color: #9A8000; font-weight: 600; margin-bottom: 4px;
}

/* Layer badges */
.layer-badge { display: inline-block; border-radius: 20px; padding: 2px 10px;
    font-size: 0.68rem; font-weight: 600; letter-spacing: 1px; text-transform: uppercase; margin-bottom: 6px; }
.layer-1 { background: #FFFDE0; color: #7A6000; }
.layer-2 { background: #FFF9C0; color: #5C4A00; }
.layer-3 { background: #FFF7ED; color: #9A3412; }

/* Risk badges */
.risk-badge { display: inline-block; border-radius: 20px; padding: 1px 9px;
    font-size: 0.65rem; font-weight: 600; text-transform: uppercase; margin-left: 6px; }
.risk-low  { background: #DCFCE7; color: #166534; }
.risk-med  { background: #FEF3C7; color: #92400E; }
.risk-high { background: #FEE2E2; color: #991B1B; }

/* Org type tag */
.org-type { display: inline-block; background: #F1F5F9; color: #475569;
    font-size: 0.65rem; font-weight: 500; padding: 1px 8px;
    border-radius: 6px; margin-left: 6px; border: 1px solid #E2E8F0; }

.sw-card {
    background: #f8fafc; border: 1px solid #e2e8f0;
    border-radius: 14px; padding: 18px 20px; margin-bottom: 12px;
}
.sw-card.l1 { border-left: 4px solid #E8C800; }
.sw-card.l2 { border-left: 4px solid #D4A800; }
.sw-card.l3 { border-left: 4px solid #B88A00; }
.sw-name  { font-size: 1rem; font-weight: 600; color: #0D1B2A; margin-bottom: 4px; }
.sw-desc  { font-size: 0.82rem; color: #475569; line-height: 1.5; margin-bottom: 6px; }
.sw-detail{ font-size: 0.78rem; color: #64748b; margin-bottom: 2px; }
.sw-detail strong { color: #334155; }

.score-bar {
    display: flex; align-items: center; gap: 6px; margin-top: 8px;
}
.score-label { font-size: 0.7rem; color: #94a3b8; }
.score-val   { font-size: 0.78rem; font-weight: 600; color: #0D1B2A; }
.score-pill  {
    height: 6px; border-radius: 3px; flex: 1;
    background: linear-gradient(90deg, #7C3AED, #C4B5FD);
}
.price-pill {
    display: inline-flex; align-items: center; gap: 4px;
    border-radius: 100px; padding: 2px 9px;
    font-size: 10px; font-weight: 600; margin-left: 4px;
}
.price-free { background: #ECFDF5; color: #059669; border: 0.5px solid #A7F3D0; }
.price-paid { background: #FFFBEB; color: #D97706; border: 0.5px solid #FDE68A; }
.price-inst { background: #FFFDE0; color: #7A6000; border: 0.5px solid #E8D800; }
.score-track { flex: 1; height: 5px; background: #F5E88A; border-radius: 3px; overflow: hidden; }
.score-fill  { height: 100%; border-radius: 3px; }
.score-num   { font-size: 11px; font-weight: 600; color: #334155; min-width: 32px; text-align: right; }
.score-lbl   { font-size: 10px; color: #94A3B8; font-weight: 500; white-space: nowrap; min-width: 56px; }
.card-divider { height: 0.5px; background: #E8D800; margin: 10px 0; opacity: 0.4; }

.pkg-card {
    background: linear-gradient(135deg, #2A2000, #4A3800);
    border-radius: 14px; padding: 18px 20px; margin-bottom: 12px; color: white;
}
.pkg-title  { font-size: 1rem; font-weight: 600; color: #C4B5FD; margin-bottom: 4px; }
.pkg-obj    { font-size: 0.8rem; color: rgba(255,255,255,0.6); margin-bottom: 8px; }
.pkg-orgs   { font-size: 0.8rem; color: rgba(255,255,255,0.85); line-height: 1.6; }
.pkg-benefit{ font-size: 0.78rem; color: #86EFAC; margin-top: 8px; }
.pkg-pos    { font-size: 0.75rem; color: #FFE500; margin-top: 4px; }

.dork-card {
    background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px;
    padding: 10px 14px; margin-bottom: 8px; display: flex; align-items: flex-start; gap: 10px;
}
.dork-num  { font-size: 0.75rem; color: #94a3b8; font-weight: 600; min-width: 20px; }
.dork-text { font-family: 'DM Mono', monospace; font-size: 0.78rem; color: #7A6000; line-height: 1.5; flex: 1; word-break: break-all; }

.stats-bar {
    background: #2A2000; border-radius: 10px; padding: 12px 20px;
    display: flex; gap: 20px; align-items: center; margin-bottom: 20px; flex-wrap: wrap;
}
.stat-item { font-size: 0.82rem; color: rgba(255,255,255,0.7); }
.stat-item span { color: #FFE500; font-weight: 600; }

.aria-divider { height: 1px; background: #e2e8f0; margin: 24px 0; }
.layer-legend { display: flex; gap: 16px; margin-bottom: 16px; flex-wrap: wrap; }
.legend-item  { display: flex; align-items: center; gap: 6px; font-size: 0.78rem; color: #475569; }
.legend-dot   { width: 10px; height: 10px; border-radius: 50%; }

/* Streamlit overrides */
.stButton > button {
    background: linear-gradient(135deg, #7A6000, #E8C800) !important;
    color: #1A1A00 !important; border: none !important; border-radius: 10px !important;
    padding: 0.6rem 2rem !important; font-family: 'Sora', sans-serif !important;
    font-weight: 600 !important; font-size: 0.95rem !important; width: 100% !important;
}
.stButton > button:hover { opacity: 0.88 !important; }
.stTextArea textarea {
    border-radius: 10px !important; border: 1px solid #cbd5e1 !important;
    font-family: 'Sora', sans-serif !important; font-size: 0.9rem !important;
}
.stFileUploader { border: 1.5px dashed #94a3b8 !important; border-radius: 12px !important; padding: 8px !important; }
.stTextInput input { border-radius: 10px !important; border: 1px solid #cbd5e1 !important; }
.stDownloadButton > button {
    background: #D4A800 !important; color: #ffffff !important; border: none !important;
    border-radius: 10px !important; font-family: 'Sora', sans-serif !important;
    font-weight: 500 !important; width: 100% !important;
}
</style>
""", unsafe_allow_html=True)


# ── Helpers ──────────────────────────────────────────────────────────
def extract_text(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith('.txt'):
        return uploaded_file.read().decode('utf-8', errors='ignore')
    elif name.endswith('.pdf'):
        try:
            import pypdf
            reader = pypdf.PdfReader(uploaded_file)
            return '\n'.join(p.extract_text() or '' for p in reader.pages)
        except Exception:
            return ""
    elif name.endswith('.docx'):
        try:
            import docx
            doc = docx.Document(io.BytesIO(uploaded_file.read()))
            return '\n'.join(p.text for p in doc.paragraphs)
        except Exception:
            return ""
    return ""


# ── SYSTEM PROMPT (oculto) ────────────────────────────────────────────
SYSTEM_PROMPT = """Actúa como un consultor senior en estrategia institucional y posicionamiento académico internacional (nivel McKinsey/Bain/BCG) especializado en:
- Vinculación académica internacional
- Desarrollo de ecosistemas universitarios premium
- Posicionamiento competitivo de programas de posgrado
- Academic partnership intelligence
- Membership ecosystem strategy
- Benchmarking universitario internacional
- Desarrollo de portafolios institucionales de alto valor percibido

Tu tarea es diseñar un portafolio estratégico de asociaciones internacionales, sociedades científicas, ligas académicas, royal colleges, federaciones, research networks, organizaciones profesionales, academic consortiums, charter institutions, y cuerpos académicos internacionales alineados al dossier académico.

MODELO DE TRES CAPAS (OBLIGATORIO):
CAPA 1 — CORE DEL DOSSIER: Organizaciones directamente relacionadas con la especialidad central.
CAPA 2 — DISCIPLINAS RELACIONADAS: Organizaciones vinculadas a áreas complementarias, técnicas, científicas o interdisciplinarias.
CAPA 3 — ECOSISTEMA ESTRATÉGICO: Entidades que fortalecen liderazgo, investigación, networking, gestión, innovación, salud pública, tecnología, publicación científica, acreditación, empleabilidad o posicionamiento internacional.

FILTROS DE INCLUSIÓN (todos obligatorios):
✔ Organización académica o profesional verificable con sitio web oficial vigente
✔ Actividad reciente verificable (webinars, journals, congresos, publicaciones, eventos)
✔ Reconocimiento internacional demostrable
✔ Membership real verificable (institucional, académica, estudiantil, profesional, research, partnership, chapter, affiliate o acceso educativo)
✔ Al menos un modelo accesible: gratuito, freemium, student membership, academic discount, institutional partnership, subsidized access o educational access
✔ Valor tangible en: networking, investigación, journals, webinars, certificaciones, empleabilidad, liderazgo, acceso científico o desarrollo profesional
✔ Mapeo a al menos un módulo del dossier

FILTROS DE EXCLUSIÓN:
✗ Organizaciones inactivas o sin evidencia de actividad reciente
✗ Memberships puramente comerciales sin valor académico real
✗ Asociaciones sin comunidad, eventos ni publicaciones
✗ Organizaciones de baja legitimidad académica o comportamiento depredador (pay-to-join sin valor)
✗ Directorios o agregadores sin comunidad institucional
✗ Organizaciones excesivamente locales sin impacto internacional
✗ Organizaciones político-partidistas o de reputación dudosa

PRIORIZACIÓN ESTRATÉGICA — dar mayor prioridad a:
- Organizaciones que eleven el prestigio internacional percibido
- Journals indexados o publicaciones relevantes
- Congresos internacionales
- Networking activo y comunidades reales
- Presencia multicontinental
- Relevancia para acreditación o reputación institucional
- Difíciles de obtener para universidades promedio

CONDICIONES NO NEGOCIABLES:
- Mínimo 50 organizaciones sin duplicados
- Nunca inventar organizaciones; solo URLs oficiales verificables
- Aplicar juicio crítico real; no descripciones promocionales
- Toda organización debe mapear al dossier

Responde ÚNICAMENTE con JSON válido sin texto adicional ni bloques de código:
{
  "asociaciones": [
    {
      "nombre": "string",
      "tipo_organizacion": "string (Society | Royal College | Federation | Research Network | Consortium | Academic Association | Professional Body | Charter Institution | otro)",
      "tipo_membership": "string",
      "capa": 1,
      "tipo_acceso": "string (Free | Freemium | Student Membership | Academic Discount | Institutional Partnership | subsidized | otro)",
      "url_oficial": "string",
      "relacion_dossier": "string (máx 25 palabras + módulo relacionado)",
      "valor_estrategico_institucional": "string",
      "valor_visible_estudiante": "string",
      "potencial_networking": "string (Alto | Medio | Bajo)",
      "prestigio_internacional": "string (Alto | Medio | Bajo)",
      "riesgo_reputacional": "string (Bajo | Medio | Alto)",
      "paquete": "string (Hospitales | Medicina y Clínica | Cuidados Médicos | Cuidados Intensivos | Investigación y Publicación | Dirección y Liderazgo | Salud Pública y Políticas | Innovación y Tecnología Médica | Educación Médica | Networking Internacional)",
      "score_relevancia": 4,
      "score_prestigio": 5,
      "score_cv": 4,
      "score_networking": 3,
      "score_uso_profesional": 4,
      "score_diferenciacion": 4,
      "score_acreditacion": 3,
      "score_visibilidad": 4
    }
  ],
  "conceptos": [
    {
      "palabra_clave": "string",
      "modulo_relacionado": "string",
      "tecnologias_asociadas": "string",
      "metodos_frameworks": "string",
      "areas_aplicacion": "string",
      "tipo_organizaciones": "string",
      "potencial_networking": "string",
      "tendencia_academica": "string"
    }
  ],
  "google_dorks": ["dork1", "dork2"]
}

Scores son enteros 1-5. Genera mínimo 50 asociaciones y mínimo 12 conceptos clave."""


def search_with_gemini(api_key: str, dossier_text: str, user_context: str) -> dict:
    genai.configure(api_key=api_key)

    user_message = f"""Analiza el siguiente dossier académico y construye el portafolio estratégico completo de asociaciones internacionales.

DOSSIER ACADÉMICO:
{dossier_text[:6000] if dossier_text else 'No se proporcionó dossier. Usa el contexto adicional.'}

CONTEXTO ADICIONAL DEL USUARIO:
{user_context or 'Programa de posgrado en ciencias de la salud.'}

Busca activamente en internet para verificar cada organización antes de incluirla.
Asegúrate de incluir mínimo 50 organizaciones y que todas tengan URL oficial verificable.
Responde SOLO con el JSON."""

    model = genai.GenerativeModel(
        model_name="gemini-2.5-flash",
        system_instruction=SYSTEM_PROMPT,
    )
    response = model.generate_content(user_message)
    raw = response.text
    try:
        clean = re.sub(r'```[a-z]*', '', raw)
        clean = clean.replace('`', '').strip()
        start = clean.index('{')
        end = clean.rindex('}') + 1
        return json.loads(clean[start:end])
    except Exception:
        pass
    return {"asociaciones": [], "conceptos": [], "google_dorks": [], "_raw": raw[:800]}


# ── Colors & constants ───────────────────────────────────────────────
CAPA_COLORS = {1: "5B21B6", 2: "1D4ED8", 3: "EA580C"}
CAPA_BG     = {1: "EDE9FE", 2: "DBEAFE", 3: "FFEDD5"}
RISK_COLORS = {"Bajo": "166534", "Medio": "92400E", "Alto": "991B1B"}
RISK_BG     = {"Bajo": "DCFCE7", "Medio": "FEF3C7", "Alto": "FEE2E2"}
NET_COLORS  = {"Alto": "065F46", "Medio": "1D4ED8", "Bajo": "6B7280"}

PAQUETES_ORDER = [
    "Hospitales", "Medicina y Clínica", "Cuidados Médicos", "Cuidados Intensivos",
    "Investigación y Publicación", "Dirección y Liderazgo", "Salud Pública y Políticas",
    "Innovación y Tecnología Médica", "Educación Médica", "Networking Internacional"
]
PKG_ICONS = {
    "Hospitales": "🏥", "Medicina y Clínica": "🩺", "Cuidados Médicos": "💊",
    "Cuidados Intensivos": "🫀", "Investigación y Publicación": "🔬",
    "Dirección y Liderazgo": "📊", "Salud Pública y Políticas": "🌎",
    "Innovación y Tecnología Médica": "💡", "Educación Médica": "🎓",
    "Networking Internacional": "🤝"
}


# ── Excel builder ─────────────────────────────────────────────────────
def _hdr(ws, row, col, value, bg="1A3C5E", fg="FFFFFF", sz=11):
    c = ws.cell(row, col, value)
    c.font      = Font(bold=True, color=fg, name="Arial", size=sz)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def _cell(ws, row, col, value, bold=False, color="000000", bg=None, halign="left", sz=10):
    c = ws.cell(row, col, value)
    c.font      = Font(bold=bold, color=color, name="Arial", size=sz)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
    thin = Side(style="thin", color="E2E8F0")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

def title_row(ws, text, cols, bg="0D1B2A"):
    ws.merge_cells(f"A1:{openpyxl.utils.get_column_letter(cols)}1")
    c = ws.cell(1, 1, text)
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

def build_excel(data: dict) -> bytes:
    import openpyxl
    wb    = openpyxl.Workbook()
    asocs = data.get("asociaciones", [])
    concs = data.get("conceptos", [])
    dorks = data.get("google_dorks", [])

    # ── HOJA 1: PORTAFOLIO ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Portafolio Estratégico"
    title_row(ws1, "PORTAFOLIO ESTRATÉGICO DE MEMBRESÍAS — ARIA Membresías", 12, "2A2000")

    h1 = ["#", "Nombre", "Tipo de Organización", "Tipo de Membership",
          "Capa", "Tipo de Acceso", "Link Oficial",
          "Relación con el Dossier", "Valor Estratégico Institucional",
          "Valor Visible para el Estudiante", "Networking", "Prestigio", "Riesgo Reputacional"]
    w1 = [4, 26, 22, 22, 7, 20, 32, 34, 34, 30, 10, 10, 14]
    for i, (h, w) in enumerate(zip(h1, w1), 1):
        _hdr(ws1, 2, i, h, bg="3D3500")
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws1.row_dimensions[2].height = 34

    for idx, a in enumerate(asocs, 1):
        row = idx + 2
        capa = a.get("capa", 1)
        risk = a.get("riesgo_reputacional", "Bajo")
        net  = a.get("potencial_networking", "Medio")
        pre  = a.get("prestigio_internacional", "Medio")
        alt  = "F8F5FF" if idx % 2 == 0 else "FFFFFF"

        vals = [
            idx,
            a.get("nombre", ""),
            a.get("tipo_organizacion", ""),
            a.get("tipo_membership", ""),
            f"Capa {capa}",
            a.get("tipo_acceso", ""),
            a.get("url_oficial", ""),
            a.get("relacion_dossier", ""),
            a.get("valor_estrategico_institucional", ""),
            a.get("valor_visible_estudiante", ""),
            net,
            pre,
            risk,
        ]
        for col, val in enumerate(vals, 1):
            cell_bg = CAPA_BG.get(capa, alt) if col == 5 else alt
            cell_fg = CAPA_COLORS.get(capa, "000000") if col == 5 else "334155"
            is_link = col == 7
            is_risk = col == 13
            is_net  = col == 11
            is_pre  = col == 12

            if is_risk:
                cell_bg = RISK_BG.get(risk, alt)
                cell_fg = RISK_COLORS.get(risk, "000000")
            elif is_net or is_pre:
                level = val
                cell_bg = {"Alto": "DCFCE7", "Medio": "FEF3C7", "Bajo": "F1F5F9"}.get(level, alt)
                cell_fg = {"Alto": "065F46", "Medio": "92400E", "Bajo": "6B7280"}.get(level, "334155")

            _cell(ws1, row, col, val,
                  bold=(col in (2, 5)),
                  color="1A6FA8" if is_link else (cell_fg if col in (5, 11, 12, 13) else "334155"),
                  bg=cell_bg,
                  halign="center" if col in (1, 5, 11, 12, 13) else "left")
        ws1.row_dimensions[row].height = 55

    ws1.freeze_panes = "A3"
    ws1.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(h1))}{len(asocs)+2}"

    # Leyenda
    leg = len(asocs) + 4
    ws1.cell(leg, 1, "LEYENDA DE CAPAS").font = Font(bold=True, name="Arial", size=10)
    for capa, label in [(1,"Capa 1 — Core del Dossier"),(2,"Capa 2 — Disciplinas Relacionadas"),(3,"Capa 3 — Ecosistema Estratégico")]:
        leg += 1
        c = ws1.cell(leg, 1, label)
        c.font = Font(bold=True, color=CAPA_COLORS[capa], name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor=CAPA_BG[capa])

    # ── HOJA 2: PAQUETES ────────────────────────────────────────────
    ws2 = wb.create_sheet("Paquetes Estratégicos")
    title_row(ws2, "PAQUETES ESTRATÉGICOS POR CONTEXTO", 6, "3D3500")
    h2 = ["Paquete", "Objetivo Estratégico", "Organizaciones", "Beneficio para el Estudiante",
          "Valor Competitivo Institucional", "Tipo de Posicionamiento"]
    w2 = [22, 38, 50, 36, 36, 28]
    for i, (h, w) in enumerate(zip(h2, w2), 1):
        _hdr(ws2, 2, i, h, bg="3D3500")
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.row_dimensions[2].height = 30

    paquetes: dict = {}
    for a in asocs:
        pkg = a.get("paquete", "Otros")
        paquetes.setdefault(pkg, []).append(a)

    pkg_bgs = ["3D3500","1D4ED8","166534","7C3AED","9A3412","065F46","0D1B2A","1A3C5E","5B21B6","3730A3"]
    row2 = 3
    for pi, pname in enumerate(PAQUETES_ORDER):
        sws = paquetes.get(pname, [])
        if not sws:
            continue
        bg_col = pkg_bgs[pi % len(pkg_bgs)]
        names  = ", ".join(s.get("nombre","") for s in sws[:15])
        vals2  = [
            f"{PKG_ICONS.get(pname,'📦')} {pname}",
            f"Conectar al estudiante con ecosistemas reales de {pname.lower()} a nivel internacional.",
            names,
            f"Acceso a comunidad profesional, journals y eventos en {pname.lower()}.",
            f"Diferenciación visible frente a programas sin vinculación en {pname.lower()}.",
            "Prestigio internacional + Networking + Empleabilidad",
        ]
        for col, val in enumerate(vals2, 1):
            c = ws2.cell(row2, col, val)
            c.font      = Font(bold=(col==1), color="FFFFFF" if col==1 else "0D1B2A", name="Arial", size=10)
            c.fill      = PatternFill("solid", fgColor=bg_col if col==1 else ("F8F5FF" if row2%2==0 else "FFFFFF"))
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin = Side(style="thin", color="E2E8F0")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws2.row_dimensions[row2].height = 65
        row2 += 1
    ws2.freeze_panes = "A3"

    # ── HOJA 3: CONCEPTOS ───────────────────────────────────────────
    ws3 = wb.create_sheet("Mapa de Conceptos")
    title_row(ws3, "MAPA DE CONCEPTOS DEL DOSSIER — ECOSISTEMA DE ASOCIACIONES", 8, "3D3500")
    h3 = ["Concepto / Palabra Clave", "Módulo Relacionado", "Tecnologías Asociadas",
          "Métodos / Frameworks", "Áreas de Aplicación", "Tipo de Organizaciones",
          "Potencial de Networking", "Tendencia Académica"]
    w3 = [28, 24, 30, 28, 28, 28, 20, 28]
    for i, (h, w) in enumerate(zip(h3, w3), 1):
        _hdr(ws3, 2, i, h, bg="3D3500")
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws3.row_dimensions[2].height = 30

    for idx, c in enumerate(concs, 1):
        row3 = idx + 2
        bg3  = "F5F3FF" if idx%2==0 else "FFFFFF"
        keys = ["palabra_clave","modulo_relacionado","tecnologias_asociadas",
                "metodos_frameworks","areas_aplicacion","tipo_organizaciones",
                "potencial_networking","tendencia_academica"]
        for col, key in enumerate(keys, 1):
            _cell(ws3, row3, col, c.get(key,""), bg=bg3,
                  bold=(col==1), color="5B21B6" if col==1 else "334155")
        ws3.row_dimensions[row3].height = 42
    ws3.freeze_panes = "A3"

    # ── HOJA 4: SCORING ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Scoring Estratégico")
    title_row(ws4, "SCORING ESTRATÉGICO PONDERADO — EVALUACIÓN COMPARATIVA", 10, "2A2000")

    # Ponderaciones: Relevancia 25%, Prestigio 25%, CV 20%, Networking 15%, Diferenciación 15%
    h4 = ["#", "Nombre", "Capa",
          "Relevancia\nAcadémica\n(×0.25)",
          "Prestigio\nInternacional\n(×0.25)",
          "Valor CV\n(×0.20)",
          "Networking\n(×0.15)",
          "Uso\nProfesional",
          "Diferenciación\n(×0.15)",
          "Acreditación",
          "Visibilidad",
          "SCORE\nPONDERADO"]
    w4 = [4, 30, 8, 14, 15, 12, 12, 10, 14, 12, 12, 14]
    for i, (h, w) in enumerate(zip(h4, w4), 1):
        _hdr(ws4, 2, i, h, bg="3D3500")
        ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws4.row_dimensions[2].height = 42

    # Ponderación note
    ws4.merge_cells("A3:L3")
    note = ws4.cell(3, 1, "⚡ Fórmula ponderada: Relevancia×25% + Prestigio×25% + CV×20% + Networking×15% + Diferenciación×15%  |  Escala: 1–5 por criterio  |  Score máx: 5.0")
    note.font      = Font(italic=True, color="5B21B6", name="Arial", size=9)
    note.fill      = PatternFill("solid", fgColor="EDE9FE")
    note.alignment = Alignment(horizontal="left", vertical="center")
    ws4.row_dimensions[3].height = 20

    for idx, a in enumerate(asocs, 1):
        row4 = idx + 3
        capa = a.get("capa", 1)
        alt4 = "F8F5FF" if idx%2==0 else "FFFFFF"

        _cell(ws4, row4, 1, idx, halign="center", bg=alt4)
        _cell(ws4, row4, 2, a.get("nombre",""), bold=True,
              color=CAPA_COLORS.get(capa,"000000"), bg=CAPA_BG.get(capa, alt4))
        _cell(ws4, row4, 3, f"C{capa}", halign="center",
              color=CAPA_COLORS.get(capa,"000000"), bg=CAPA_BG.get(capa, alt4), bold=True)

        score_keys = ["score_relevancia","score_prestigio","score_cv","score_networking",
                      "score_uso_profesional","score_diferenciacion","score_acreditacion","score_visibilidad"]
        for col, key in zip(range(4, 12), score_keys):
            val = a.get(key, 3)
            _cell(ws4, row4, col, val, halign="center", bg=alt4)

        # Weighted score formula: D*0.25 + E*0.25 + F*0.20 + G*0.15 + I*0.15
        r = row4
        formula = f"=ROUND(D{r}*0.25+E{r}*0.25+F{r}*0.20+G{r}*0.15+I{r}*0.15,2)"
        tc = ws4.cell(row4, 12, formula)
        tc.font      = Font(bold=True, name="Arial", size=10)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="E2E8F0")
        tc.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        tc.fill      = PatternFill("solid", fgColor=alt4)
        ws4.row_dimensions[row4].height = 22

    # Color scale on score columns
    last_row = len(asocs) + 3
    for col_letter in ["D","E","F","G","H","I","J","K"]:
        ws4.conditional_formatting.add(
            f"{col_letter}4:{col_letter}{last_row}",
            ColorScaleRule(start_type="min", start_color="FFCCCC",
                           mid_type="percentile", mid_value=50, mid_color="FFFFCC",
                           end_type="max", end_color="CCFFCC")
        )
    ws4.conditional_formatting.add(
        f"L4:L{last_row}",
        ColorScaleRule(start_type="min", start_color="FF4444",
                       mid_type="percentile", mid_value=50, mid_color="FFBB33",
                       end_type="max", end_color="00C851")
    )
    ws4.freeze_panes = "A4"

    # ── HOJA 5: GOOGLE DORKS ────────────────────────────────────────
    ws5 = wb.create_sheet("Google Dorks")
    title_row(ws5, "GOOGLE DORKS — VERIFICACIÓN Y BÚSQUEDA AVANZADA DE ASOCIACIONES", 3, "3D3500")
    for i, (h, w) in enumerate(zip(["#","Google Dork","Uso / Verificación Sugerida"],[6,80,40]),1):
        _hdr(ws5, 2, i, h, bg="3D3500")
        ws5.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws5.row_dimensions[2].height = 26

    for idx, dork in enumerate(dorks, 1):
        row5 = idx + 2
        bg5  = "F5F3FF" if idx%2==0 else "FFFFFF"
        _cell(ws5, row5, 1, idx, halign="center", bg=bg5)
        c = ws5.cell(row5, 2, dork)
        c.font      = Font(name="Courier New", size=10, color="5B21B6")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="E2E8F0")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        c.fill   = PatternFill("solid", fgColor=bg5)
        _cell(ws5, row5, 3, "", bg=bg5)
        ws5.row_dimensions[row5].height = 22
    ws5.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── UI ────────────────────────────────────────────────────────────────
st.markdown("""
<div class="aria-header">
  <div class="aria-eyebrow">&#9670; Herramienta Académica</div>
  <h1 class="aria-title">ARIA <em>Membresías</em></h1>
  <p class="aria-sub">Portafolio estratégico de membresías y ecosistemas académicos internacionales</p>
</div>
""", unsafe_allow_html=True)

# Cargar API Key desde Secrets automáticamente
if "gemini_key" not in st.session_state:
    secret_key = st.secrets.get("GEMINI_API_KEY", "")
    if secret_key:
        st.session_state["gemini_key"] = secret_key

# Mostrar campo manual solo si no hay key precargada
if not st.session_state.get("gemini_key"):
    with st.expander("🔑 Configurar API Key de Gemini", expanded=True):
        api_key_input = st.text_input(
            "Google Gemini API Key", type="password", placeholder="AIza...",
            help="Obtén tu clave gratuita en https://aistudio.google.com/apikey"
        )
        if api_key_input:
            st.session_state["gemini_key"] = api_key_input
            st.success("✅ API Key guardada para esta sesión")

api_key = st.session_state.get("gemini_key", "")

st.markdown('<div class="section-label">01 — Dossier académico (PDF, TXT o DOCX)</div>', unsafe_allow_html=True)
uploaded = st.file_uploader(
    "Dossier", type=["pdf","txt","docx"], label_visibility="collapsed"
)

st.markdown('<div style="height:10px"></div>', unsafe_allow_html=True)
st.markdown('<div class="section-label">02 — Contexto adicional (opcional)</div>', unsafe_allow_html=True)
user_context = st.text_area(
    "Contexto", label_visibility="collapsed",
    placeholder="Ej: Maestría en Administración de Hospitales, enfocada en directivos de clínicas privadas en México...",
    height=90
)

if st.button("🏆  Generar portafolio de membresías", use_container_width=True):
    if not api_key:
        st.error("⚠️ Ingresa tu API Key de Gemini.")
    elif not uploaded and not user_context:
        st.warning("Sube el dossier o describe el contexto del programa.")
    else:
        dossier_text = extract_text(uploaded) if uploaded else ""
        with st.spinner("Analizando dossier y buscando asociaciones con Gemini... (30–60 seg)"):
            try:
                result = search_with_gemini(api_key, dossier_text, user_context)
            except Exception as e:
                st.error(f"Error al consultar Gemini: {e}")
                st.stop()

        if not result.get("asociaciones"):
            st.warning("No se obtuvieron resultados. Intenta con otro dossier o contexto.")
            if result.get("_raw"):
                with st.expander("Respuesta cruda"):
                    st.text(result["_raw"])
            st.stop()

        st.session_state["result_asoc"] = result

# ── Results ───────────────────────────────────────────────────────────
if "result_asoc" in st.session_state:
    result = st.session_state["result_asoc"]
    asocs  = result.get("asociaciones", [])
    concs  = result.get("conceptos", [])
    dorks  = result.get("google_dorks", [])

    c1 = [a for a in asocs if a.get("capa")==1]
    c2 = [a for a in asocs if a.get("capa")==2]
    c3 = [a for a in asocs if a.get("capa")==3]

    st.markdown('<div class="aria-divider"></div>', unsafe_allow_html=True)

    col_s, col_d = st.columns([2,1])
    with col_s:
        st.markdown(f"""
        <div class="stats-bar">
          <div class="stat-item"><span>{len(asocs)}</span> asociaciones</div>
          <div class="stat-item"><span>{len(c1)}</span> Capa 1</div>
          <div class="stat-item"><span>{len(c2)}</span> Capa 2</div>
          <div class="stat-item"><span>{len(c3)}</span> Capa 3</div>
          <div class="stat-item"><span>{len(concs)}</span> conceptos</div>
        </div>""", unsafe_allow_html=True)
    with col_d:
        excel_bytes = build_excel(result)
        fname = f"ARIA_Asociaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button("⬇️ Descargar Excel (5 hojas)", data=excel_bytes,
                           file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

    st.markdown("""
    <div class="layer-legend">
      <div class="legend-item"><div class="legend-dot" style="background:#E8C800"></div>Capa 1 — Core del Dossier</div>
      <div class="legend-item"><div class="legend-dot" style="background:#D4A800"></div>Capa 2 — Disciplinas Relacionadas</div>
      <div class="legend-item"><div class="legend-dot" style="background:#B88A00"></div>Capa 3 — Ecosistema Estratégico</div>
    </div>""", unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs([
        f"🏛 Portafolio ({len(asocs)})",
        f"🗂 Paquetes",
        f"🔑 Conceptos ({len(concs)})",
        f"🔎 Google Dorks ({len(dorks)})"
    ])

    with tab1:
        capa_labels = {1:"Capa 1 — Core", 2:"Capa 2 — Relacionadas", 3:"Capa 3 — Ecosistema"}
        badge_cls   = {1:"layer-1", 2:"layer-2", 3:"layer-3"}
        card_cls    = {1:"l1", 2:"l2", 3:"l3"}
        risk_cls    = {"Bajo":"risk-low","Medio":"risk-med","Alto":"risk-high"}

        for a in asocs:
            capa=a.get("capa",1); risk=a.get("riesgo_reputacional","Bajo")
            net=a.get("potencial_networking","Medio"); pre=a.get("prestigio_internacional","Medio")
            url=a.get("url_oficial",""); tipo=a.get("tipo_organizacion","")
            s_rel=a.get("score_relevancia",3); s_pre=a.get("score_prestigio",3)
            s_cv=a.get("score_cv",3); s_net=a.get("score_networking",3); s_dif=a.get("score_diferenciacion",3)
            ponderado=round(s_rel*0.25+s_pre*0.25+s_cv*0.20+s_net*0.15+s_dif*0.15,2)
            pond_pct=int(ponderado/5*100)
            rel_pct=int(s_rel/5*100); pre_pct=int(s_pre/5*100); cv_pct=int(s_cv/5*100)
            bar_col="linear-gradient(90deg,#E8C800,#059669)" if ponderado>=4 else ("linear-gradient(90deg,#E8C800,#D4A800)" if ponderado>=3 else "#D97706")
            acceso=a.get("tipo_acceso","").lower()
            if any(x in acceso for x in ["grat","free","sin cost","student member"]):
                pcls,picon,plabel="price-free","ti-circle-check","Membership gratuito"
            elif any(x in acceso for x in ["descuento","discount","reduced","academic discount"]):
                pcls,picon,plabel="price-paid","ti-tag","Descuento académico"
            else:
                pcls,picon,plabel="price-inst","ti-building","Membership institucional"
            link_html=f'<a href="{url}" target="_blank" style="color:#7A6000;font-size:12px;font-weight:500;">Sitio oficial</a>' if url else ""
            capa_l=capa_labels.get(capa,""); badge_c=badge_cls.get(capa,"layer-1"); card_c=card_cls.get(capa,"l1"); risk_c=risk_cls.get(risk,"risk-low")
            nombre=a.get("nombre",""); val_inst=a.get("valor_estrategico_institucional",""); rel_dos=a.get("relacion_dossier","")
            val_est=a.get("valor_visible_estudiante",""); memb=a.get("tipo_membership",""); acc=a.get("tipo_acceso","")
            st.markdown(f"""
            <div class="sw-card {card_c}">
              <div class="sw-name">{nombre} <span class="org-type">{tipo}</span> <span class="risk-badge {risk_c}">Riesgo {risk}</span></div>
              <div style="display:flex;align-items:center;flex-wrap:wrap;gap:4px;margin-bottom:6px;">
                <span class="layer-badge {badge_c}">{capa_l}</span>
                <span class="price-pill {pcls}"><i class="ti {picon}" aria-hidden="true" style="font-size:11px;"></i> {plabel}</span>
              </div>
              <div class="sw-desc">{val_inst}</div>
              <div class="sw-detail"><strong>Módulo:</strong> {rel_dos}</div>
              <div class="sw-detail"><strong>Para el estudiante:</strong> {val_est}</div>
              <div class="sw-detail"><strong>Membership:</strong> {memb} · <strong>Acceso:</strong> {acc}</div>
              <div class="sw-detail">Networking: <strong>{net}</strong> &nbsp;|&nbsp; Prestigio: <strong>{pre}</strong></div>
              <div class="card-divider"></div>
              <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px;">
                <span class="score-lbl">Score ponderado</span>
                <div class="score-track"><div class="score-fill" style="width:{pond_pct}%;background:{bar_col};"></div></div>
                <span class="score-num">{ponderado}/5</span>
              </div>
              <div style="display:flex;gap:10px;margin-top:4px;">
                <div style="display:flex;align-items:center;gap:6px;flex:1;">
                  <span class="score-lbl">Relevancia</span>
                  <div class="score-track"><div class="score-fill" style="width:{rel_pct}%;background:#E8C800;"></div></div>
                  <span class="score-num" style="font-size:10px;">{s_rel}/5</span>
                </div>
                <div style="display:flex;align-items:center;gap:6px;flex:1;">
                  <span class="score-lbl">Prestigio</span>
                  <div class="score-track"><div class="score-fill" style="width:{pre_pct}%;background:#059669;"></div></div>
                  <span class="score-num" style="font-size:10px;">{s_pre}/5</span>
                </div>
                <div style="display:flex;align-items:center;gap:6px;flex:1;">
                  <span class="score-lbl">Valor CV</span>
                  <div class="score-track"><div class="score-fill" style="width:{cv_pct}%;background:#1A7FE8;"></div></div>
                  <span class="score-num" style="font-size:10px;">{s_cv}/5</span>
                </div>
              </div>
              <div style="margin-top:10px;">{link_html}</div>
            </div>""", unsafe_allow_html=True)

    with tab2:
        paquetes: dict = {}
        for a in asocs:
            pkg = a.get("paquete","Otros")
            paquetes.setdefault(pkg, []).append(a)

        for pname in PAQUETES_ORDER:
            sws = paquetes.get(pname, [])
            if not sws:
                continue
            icon  = PKG_ICONS.get(pname,"📦")
            names = " · ".join(s.get("nombre","") for s in sws)
            st.markdown(f"""
            <div class="pkg-card">
              <div class="pkg-title">{icon} {pname} <span style="font-size:0.75rem;opacity:.6;">({len(sws)} organizaciones)</span></div>
              <div class="pkg-obj">Conectar al estudiante con ecosistemas reales de {pname.lower()} a nivel internacional.</div>
              <div class="pkg-orgs">{names}</div>
              <div class="pkg-benefit">✓ Acceso a comunidad profesional, journals y eventos internacionales.</div>
              <div class="pkg-pos">🎯 Posicionamiento: Prestigio internacional · Networking · Empleabilidad</div>
            </div>""", unsafe_allow_html=True)

    with tab3:
        for concept in concs:
            with st.expander(f"🔑 {concept.get('palabra_clave','')} — {concept.get('modulo_relacionado','')}"):
                st.markdown(f"**Tecnologías asociadas:** {concept.get('tecnologias_asociadas','')}")
                st.markdown(f"**Métodos / Frameworks:** {concept.get('metodos_frameworks','')}")
                st.markdown(f"**Áreas de aplicación:** {concept.get('areas_aplicacion','')}")
                st.markdown(f"**Tipo de organizaciones:** {concept.get('tipo_organizaciones','')}")
                st.markdown(f"**Potencial de networking:** {concept.get('potencial_networking','')}")
                st.markdown(f"**Tendencia académica:** {concept.get('tendencia_academica','')}")

    with tab4:
        st.caption("Copia cualquier dork para verificar asociaciones directamente en Google")
        for i, dork in enumerate(dorks, 1):
            st.markdown(f"""
            <div class="dork-card">
              <span class="dork-num">{i}</span>
              <span class="dork-text">{dork}</span>
            </div>""", unsafe_allow_html=True)

    st.markdown(
        '<p style="text-align:center;font-size:0.72rem;color:#94a3b8;margin-top:32px;">'
        'ARIA Membresías · Portafolio Estratégico · Powered by Google Gemini 2.5 Flash</p>',
        unsafe_allow_html=True)
